#################################Autor: Jorge Eliécer Loaiza Muñoz###################################################

# Este script permite validar una serie de documentos en la plataforma SOFIAPLUS y
# rellena en rojo aquellos que no están registrados, y en verde los que si.
# Es necesario que el archivo esté cerrado en el momento de iniciar la validación.
# El listado de documentos debe estar en la columna "A" de la hoja de cálculo y apartir de la fila 1.
# Todos los números de documentos deben ser del mismo tipo, es decir, todos deben ser CC o TI o CE o PEP, no pueden
# estar combinados.
#Se utiliza la librería Tkinter para la GUI.

#from browser import document, window
#from browser.html import *
import time
from selenium import webdriver
from selenium.webdriver.support.ui import Select
import openpyxl
from openpyxl.styles import PatternFill
from tkinter import *
from tkinter import ttk, Button
from tkinter import filedialog, messagebox
from multiprocessing import Process
from threading import Thread  # librería para ejecutar en paralelo
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.firefox import GeckoDriverManager
from webdriver_manager.opera import OperaDriverManager


opcion_documento = 0  # esta variable la declaro global para no retornarla y saber que tipo de documento seleccionaron para trabajar
path_name = ""  # declaro esta variable global para no retornarla en la función


def open_file():  # esta función abre el archivo a trabajar
    global path_name  # variable donde se guardará la ruta del archivo
    # esta línea abre la ventana para buscar el archivo y guarda en la variable la ruta del archivo seleccionado
    path_name = filedialog.askopenfilename(initialdir="/", title="Abrir archivo",
                                           filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))

    Label(root, text=path_name).place(x=75, y=45)  # muestra sobre la ventana el path del archivo


def iniciar_validacion():  # esta función realiza el proceso de comprobación de registro de los documentos
    try:

        global path_name  # variable que contiene la ruta del archivo
        # cargo el libro donde están los números de identificación
        wb = openpyxl.load_workbook(path_name)
        # se ubica en la hoja activa del libro donde deben estar los datos a verificar
        sheet = wb.active
        # captura el número de la fila que contiene el último dato, es decir, la cantidad de filas con datos
        ultima_fila_con_datos = sheet.max_row
        print("Cantidad de datos a validar ",ultima_fila_con_datos)

        #driver = webdriver.Chrome(executable_path=r'C:\chromedriver.exe')

        #esta línea detecta automáticamente el "chromedriver.exe" y lo instala si no está instalado
        driver = webdriver.Chrome(executable_path=ChromeDriverManager().install())  # para google chrome
        #esta línea detecta automáticamente el "operadriver.exe.exe" y lo instala si no está instalado
        #driver = webdriver.Opera(executable_path=OperaDriverManager().install())#para Opera
        #esta línea detecta automáticamente el "GeckoDriver.exe" y lo instala si no está instalado
        #driver = webdriver.Firefox(executable_path=GeckoDriverManager().install())#para mozilla firefox

        driver.maximize_window()

        fill_pattern_verde = PatternFill(patternType='solid', fgColor='0099CC00')# esta línea alista el color verde
        fill_pattern_rojo = PatternFill(patternType='solid', fgColor='00FF0000') # esta línea alista el color rojo

        # este ciclo recorre todas las filas que contienen cada uno de los documentos a validar
        for i in range(ultima_fila_con_datos):
            driver.get('http://oferta.senasofiaplus.edu.co/sofia-oferta/inicio-sofia-plus.html')  # abre la página de SOFIA
            time.sleep(1)  # le coloco este retraso de un segundo porque en unas pruebas no estaba cerrando la ventana.Tal vez esto sucedía
            # por lo rápido que se ejecuta el código
            driver.find_element_by_xpath('//*[@id="area_trabajo_divNoticiaPrincipal"]/div/span[1]/img').click()  # cierra el aviso inicial
            driver.find_element_by_xpath('//*[@id="registro"]')  # hace click en el botón "Registrarse"
            driver.get('http://oferta.senasofiaplus.edu.co/sofia-oferta/registro.html')  # se ubica en la página del registro
            # esta línea permite cambiar al frame donde está el formulario para verificar si se está registrado
            time.sleep(1)
            driver.switch_to.frame(driver.find_element_by_xpath('//*[@id="modal-content"]/iframe'))

            # Esta línea ubica la lista desplegable de los tipo de documentos de identidad
            element = driver.find_element_by_xpath('//*[@id="s1"]/select')
            dropdown = Select(element)
            if opcion_documento == 1:
                # esta línea selecciona el elemento de la lista desplegable "Cédula de Ciudadanía"
                dropdown.select_by_visible_text('Cédula de Ciudadanía')
            elif opcion_documento == 2:
                # esta línea selecciona el elemento de la lista desplegable "Tarjeta de Identidad"
                dropdown.select_by_visible_text('Tarjeta de Identidad')
            elif opcion_documento == 3:
                # esta línea selecciona el elemento de la lista desplegable "Cédula de Extranjeria"
                dropdown.select_by_visible_text('Cédula de Extranjeria')
            else:
                # esta línea selecciona el elemento de la lista desplegable "PEP"
                dropdown.select_by_visible_text('PEP')

            # esta línea me ubica en el cuadro de texto donde se debe ingresar el documento y borra su contenido
            driver.find_element_by_xpath('//*[@id="validar"]/label[2]/div[2]/input').clear()
            # esta línea me ubica en el cuadro de texto donde se debe ingresar el documento e ingresa un documento
            driver.find_element_by_xpath('//*[@id="validar"]/label[2]/div[2]/input').send_keys(sheet.cell(row=i + 1, column=1).value)

            # esta línea hace clic en el botón "validar"
            driver.find_element_by_xpath('//*[@id="validar"]/div[4]/button').click()

            # Espera tres segundos antes de continuar con el código de abajo. Es necesario esperar porque el aviso emergente de
            # ya registrado se puede tardar uno o dos segundos en aparecer.
            time.sleep(2)

            # Xpath del aviso emergente que indica que el documento ya está resgistrado
            display = driver.find_element_by_xpath('//*[@id="msg13"]')

            if display.is_displayed() == True:  # si es igual a True es porque apareció el aviso emergente, lo que significa que el documento ya está registrado
                sheet.cell(i + 1,1).fill = fill_pattern_verde  # rellena de verde la celda donde está el documento ya validado
                wb.save(path_name)  # guarda los cambios hechos en el libro donde están los números de identificación
                #time.sleep(10)  # espera diez segundos para que el aviso emergente se cierre
                print("Documento ya resgistrado")
            else:
                # rellena de rojo la celda del documento  si éste no está registrado
                sheet.cell(i + 1, 1).fill = fill_pattern_rojo
                # guarda los cambios hechos en el libro donde están los números de identificación en el mismo path
                wb.save(path_name)
                # botón para regresar al frame donde se introduce el documento
                #driver.find_element_by_xpath('//*[@id="registro_paso_1"]/div[2]/div[2]/button[1]').click()
                # esta línea permite cambiar al frame donde está el formulario para verificar si se está registrado
                # si no coloco esta línea, no funciona
                #driver.switch_to.frame(driver.find_element_by_xpath('//*[@id="modal-content"]/iframe'))
                print("Documento NO registrado")

        # esta línea coloca el mensaje "Los documentos han sido validados con éxito"
        Label(root, text="Documentos validados con éxito").place(x=72, y=110)
        # esta línea crea de nuevo la barra de progreso y después la detengo con .stop
        my_progress = ttk.Progressbar(root, orient=HORIZONTAL, length=250, mode='indeterminate')
        my_progress.place(x=25, y=140)  # ubica la barra de progreso en las mismas coordenadas
        my_progress.stop()  # detiene la barra de progreso
        # esta línea muestra un mensaje de finalización del proceso de validación
        messagebox.showinfo(title="Finalización de la validación",message="La validación ha finalizado. Abra el archivo para ver los resultados. Los documentos resaltados en verde ya están registrados")

        # driver.close() #cierra el navegador

    except TypeError: #este bloque de código muestra una ventana avisando que la validación finalizó.
                     #toca colocarlo asi porque a veces sale el error de tipo "TypeError"
        # esta línea coloca el mensaje "Los documentos han sido validados con éxito"
        Label(root, text="Documentos validados con éxito").place(x=72, y=110)
        # esta línea crea de nuevo la barra de progreso y después la detengo con .stop
        my_progress = ttk.Progressbar(root, orient=HORIZONTAL, length=250, mode='indeterminate')
        my_progress.place(x=25, y=140)  # ubica la barra de progreso en las mismas coordenadas
        my_progress.stop()  # detiene la barra de progreso
        # esta línea muestra un mensaje de finalización del proceso de validación
        messagebox.showinfo(title="Finalización de la validación",
                            message="La validación ha finalizado. Abra el archivo para ver los resultados. Los documentos resaltados en verde ya están registrados")


    except PermissionError: #este bloque de código lanza una ventana de aviso cuando el archivo con los documentos a validar está abierto
        # esta línea crea de nuevo la barra de progreso y después la detengo con .stop
        my_progress = ttk.Progressbar(root, orient=HORIZONTAL, length=250, mode='indeterminate')
        my_progress.place(x=25, y=140)  # ubica la barra de progreso en las mismas coordenadas
        my_progress.stop()  # detiene la barra de progreso
        messagebox.showwarning("Aviso","No se puede iniciar la validación porque el archivo que contiene los documentos está abierto. Ciérrelo e intente de nuevo")


def salir_aplicacion():  # esta función muestra una ventana para preguntar si desean salir o no de la aplicación
    respuesta = messagebox.askquestion(title="Salir", message="¿Desea salir de la aplicación?")
    if respuesta == "yes":
        root.destroy()


def barra_de_progreso():  # esta función genera la barra de progreso
    my_progress = ttk.Progressbar(root, orient=HORIZONTAL, length=250, mode='indeterminate')
    my_progress.pack(pady=10)
    my_progress.place(x=25, y=140)
    my_progress.start(2)  # este número controla la velocidad de la barra
    Label(root, text="Validando documentos...").place(x=80, y=110)


def barra_y_validacion():  # esta función ejecuta de manera paralela las funciones de la barra de progreso y la de inicio de la validación
    # pues es necesario que la barra se muestre al mismo tiempo que se comprueban los documentos
    if path_name == "":  # si este condicional es verdadero, es porque dieron clic al botón "iniciar validación" sin seleccionar un archivo
        messagebox.showwarning("Aviso", "Primero debe seleccionar un archivo válido para iniciar la validación.")
        return
    elif opcion_documento == 0:  # si esta condición es verdadera es porque no seleccionaron ningún tipo de documento
        messagebox.showwarning("Aviso", "Debe seleccionar un tipo de documento para iniciar la validación.")
        return
    else:
        t1 = Thread(target=barra_de_progreso)
        t2 = Thread(target=iniciar_validacion)
        t1.start()
        t2.start()


def acerca_de():
    messagebox.showinfo(title="Acerca de este script",
                        message="Esta aplicación fue creada por JORGE ELIÉCER LOAIZA MUÑOZ. Contacto: jelm48@misena.edu.co")


def selec():  # esta función asigna a la variable "opción_documento" el número de la opción seleccionada
    global opcion_documento
    opcion_documento = opcion.get()

#######################################...Inicio del script...##########################################################
if __name__ == '__main__':
    root = Tk()  # crea la ventana principal
    root.resizable(0,0)  # esta línea remueve el botón "maximizar" para que no redimensionen la ventana y asi no desubicar los botones
    root.iconbitmap('C:/Users/jelm4/iconoes.ico')  # ubica el ícono en la ventana principal

    barra_menu = Menu(root)  # crea la barra de menú
    root.config(menu=barra_menu, width=300, height=300)
    archivo_menu = Menu(barra_menu, tearoff=0)
    barra_menu.add_cascade(label="Acerca de...", menu=archivo_menu)
    archivo_menu.add_command(label="Acerca de",command=acerca_de)  # ejecuta la función "acerca_de" para mostrar el autor de la aplicación

    root.title("Validador SOFIA plus")  # coloca este texto en la barra de título de la ventana principal
    root.geometry("300x250")  # dimensiones de la ventana principal
    boton_abrir = Button(root, text="Cargar archivo", command=open_file).place(x=106,y=10)  # crea el botón para cargar el archivo con los documentos
    boton_validar = Button(root, text="Iniciar validación", command=barra_y_validacion).place(x=100,y=80)  # botón para iniciar verificación de documentos
    boton_salir = Button(root, text="Salir", command=salir_aplicacion).place(x=130,y=180)  # botón para salir de la ventana principal

    ##################.....Bloque de código para insertar los botones de opción....#########################################

    opcion = IntVar()  # Como StrinVar pero en entero
    Radiobutton(root, text="CC", variable=opcion, value=1, command=selec).place(x=8, y=50)
    Radiobutton(root, text="TI", variable=opcion, value=2, command=selec).place(x=8, y=70)
    Radiobutton(root, text="CE", variable=opcion, value=3, command=selec).place(x=8, y=90)
    Radiobutton(root, text="PEP", variable=opcion, value=4, command=selec).place(x=8, y=110)
    Label(root, text="Tipo de").place(x=6, y=13)
    Label(root, text="documento").place(x=6, y=31)

    root.mainloop()
